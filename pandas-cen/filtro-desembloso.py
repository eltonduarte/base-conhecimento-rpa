import os, glob
import pandas as pd
import traceback
from openpyxl import load_workbook

"""

caminho="@s6006vfs0601;prd_rpaa_001$;RPA_ROBOS;CENPES;CEN_PPC_01;Trabalho;Planilha_Planejamento_Contrato_RPA_22072021.xlsx"
caminho_save = "@s6006vfs0601;prd_rpaa_001$;RPA_ROBOS;CENPES;CEN_PPC_01;Trabalho"
nome_arquivo_principal= "Planilha_Planejamento_Contrato_RPA_22072021"
nome_arquivo_email= "SIGITEC_e_SAP_parcela_zero_13052021"


lista = [caminho,caminho_save,nome_arquivo_principal,nome_arquivo_email]
"""

def filtro(lista):

    try:

        caminho = lista[0]
        caminho_save = lista[1]
        nome_arquivo_principal = lista[2]
        nome_arquivo_email = lista[3]

        caminho= caminho.replace("@","\\\\")
        caminho_lst= caminho.split(";")

        caminho_save= caminho_save.replace("@","\\\\")
        caminho_save_lst= caminho_save.split(";")

        path_save = os.path.join(*caminho_save_lst)

        df_sigitec_sem_sap = pd.read_excel(os.path.join(*caminho_lst), sheet_name='SIGITEC sem SAP')

        df_desembolso = pd.read_excel(os.path.join(*caminho_lst), sheet_name='Desembolso')

        df_left= pd.merge(df_sigitec_sem_sap,
            df_desembolso, 
            left_on='Número do SIC/AEP / Processo',
            right_on='Número do Processo', 
            how="outer", 
            indicator=True).query('_merge=="left_only"')

        df_inner = pd.merge(df_sigitec_sem_sap,
            df_desembolso, 
            left_on='Número do SIC/AEP / Processo',
            right_on='Número do Processo', 
            how="inner", 
            indicator=True)

        df_unico_1 = df_left["Número do SIC/AEP / Processo"].drop_duplicates()

        df_unico_2 = df_inner["Número do SIC/AEP / Processo"].drop_duplicates()

        df_sigitec_sem_sap_com_des =  df_sigitec_sem_sap[ df_sigitec_sem_sap["Número do SIC/AEP / Processo"].isin(df_unico_2.tolist())]

        df_sigitec_sem_sap_sem_des =  df_sigitec_sem_sap[ df_sigitec_sem_sap["Número do SIC/AEP / Processo"].isin(df_unico_1.tolist())]

        path_save = os.path.join(*caminho_save_lst)

        book = load_workbook(path_save+"\\"+nome_arquivo_principal+".xlsx")

        writer = pd.ExcelWriter(path_save+"\\"+nome_arquivo_principal+".xlsx", engine='openpyxl') 

        writer.book = book   

        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        for chave in writer.sheets:
            if chave == "SIGITEC sem SAP com desembolso":                
                std=book[chave]                 
                book.remove(std)
                #break
                        
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)  


        df_sigitec_sem_sap_com_des.to_excel(writer,"SIGITEC sem SAP com desembolso",index=False)

        writer.save()


        writer_2 = pd.ExcelWriter(path_save+"\\"+nome_arquivo_email+".xlsx", engine='openpyxl') 

        df_sigitec_sem_sap_sem_des.to_excel(writer_2,"SIGITEC sem SAP sem desembolso",index=False)  

        writer_2.save()

        retorno = "0"
        return retorno
        
    except:
        retorno = "1"+","+str(traceback.format_exc())
        return retorno




def passa_planejamento_pluri(lista):
    try:
        caminho = lista[0]
        caminho_save = lista[1]
        nome_arquivo_principal = lista[2]
        nome_arquivo_email = lista[3]

        caminho= caminho.replace("@","\\\\")
        caminho_lst= caminho.split(";")

        caminho_save= caminho_save.replace("@","\\\\")
        caminho_save_lst= caminho_save.split(";")

        path_save = os.path.join(*caminho_save_lst)

        df_sigitec_sem_sap_des = pd.read_excel(os.path.join(*caminho_lst), sheet_name='SIGITEC sem SAP com desembolso')

        df_planejamento = pd.read_excel(os.path.join(*caminho_lst), sheet_name='Planejamento Plurianual')

        df_template_planejamento  = pd.read_excel(path_save+ r"\\" + "template_planejamento.xlsx" )

        df = df_sigitec_sem_sap_des.loc[df_sigitec_sem_sap_des.index.repeat(df_sigitec_sem_sap_des['Número de Parcelas Previstas'])]

        #lista_ind = df.index.tolist()

        lista_df = df["Número de Parcelas Previstas"].tolist()

        lista_processo = df["Número do SIC/AEP / Processo"].tolist()

        controle = len(lista_df)

        comp_num = 0
        comp_str = ""
        count = 0
        lista_mudanca = []

        for i,j,k  in zip(lista_df,range(controle),lista_processo):
            if j == 0:
                comp_num = i
                comp_str = k
            if i == comp_num and k == comp_str:
                count+=1
                lista_mudanca.append(count)        
            else:
                count = 0
                comp_num = i
                comp_str = k
                count+=1
                lista_mudanca.append(count)            

        df["Número de Parcelas Previstas"] = lista_mudanca

        df_template_planejamento["Descrição"] = df["Número do SIC/AEP / Processo"]

        df_template_planejamento["SIGITEC"] = df["Número do SIC/AEP / Processo"]

        df_template_planejamento["Parceiro Dev"] = df["Proponente"]

        df_template_planejamento["Parcela"] = df["Número de Parcelas Previstas"]

        df_planejamento = pd.concat([df_planejamento,df_template_planejamento])

        book = load_workbook(os.path.join(*caminho_lst))

        writer_2 = pd.ExcelWriter(os.path.join(*caminho_lst), engine='openpyxl')

        writer_2.book = book

        writer_2.sheets = dict((ws.title, ws) for ws in book.worksheets)

        for chave in writer_2.sheets:
            if chave == "Planejamento Plurianual":                
                std=book[chave]                 
                book.remove(std)
                break
                
        writer_2.sheets = dict((ws.title, ws) for ws in book.worksheets)
            
        df_planejamento.to_excel(writer_2, "Planejamento Plurianual",index=False)  

        writer_2.save()


        retorno = "0"+","
        return retorno
    except:
        retorno = "1"+","+str(traceback.format_exc())
        return retorno


def passa_planejamento_pluri_2(lista):
    try:
        caminho = lista[0]
        caminho_save = lista[1]
        nome_arquivo_principal = lista[2]
        nome_arquivo_email = lista[3]

        caminho= caminho.replace("@","\\\\")
        caminho_lst= caminho.split(";")

        caminho_save= caminho_save.replace("@","\\\\")
        caminho_save_lst= caminho_save.split(";")

        path_save = os.path.join(*caminho_save_lst)

        df_sigitec_e_sap = pd.read_excel(os.path.join(*caminho_lst), sheet_name='SIGITEC e SAP')

        df_planejamento = pd.read_excel(os.path.join(*caminho_lst), sheet_name='Planejamento Plurianual')

        df_template_planejamento  = pd.read_excel(path_save+ r"\\" + "template_planejamento.xlsx" )

        df_zero = df_sigitec_e_sap.loc[df_sigitec_e_sap['Número de Parcelas Previstas'] == 0]

        df_diferente_zero = df_sigitec_e_sap[df_sigitec_e_sap['Número de Parcelas Previstas'] != 0]

        df_diferente_zero = df_diferente_zero.reset_index(drop=True)

        #print(df_zero)             

        df = df_diferente_zero.loc[df_diferente_zero.index.repeat(df_diferente_zero['Número de Parcelas Previstas'])]

        #lista_ind = df.index.tolist()

        lista_df = df["Número de Parcelas Previstas"].tolist()

        lista_processo = df["Número do SIC/AEP / Processo"].tolist()

        controle = len(lista_df)

        comp_num = 0
        comp_str = ""
        count = 0
        lista_mudanca = []

        for i,j,k  in zip(lista_df,range(controle),lista_processo):
            if j == 0:
                comp_num = i
                comp_str = k
            if i == comp_num and k == comp_str:
                count+=1
                lista_mudanca.append(count)        
            else:
                count = 0
                comp_num = i
                comp_str = k
                count+=1
                lista_mudanca.append(count)            

        df["Número de Parcelas Previstas"] = lista_mudanca

        df_template_planejamento["Descrição"] = df["Número do SIC/AEP / Processo"]

        df_template_planejamento["SIGITEC"] = df["Número do SIC/AEP / Processo"]

        df_template_planejamento["Parceiro Dev"] = df["Proponente"]

        df_template_planejamento["Parcela"] = df["Número de Parcelas Previstas"]

        df_template_planejamento["SAP"] = df["Número do SAP"]

        df_planejamento = pd.concat([df_planejamento,df_template_planejamento])


        writer = pd.ExcelWriter(path_save+"\\"+nome_arquivo_email+".xlsx", engine='openpyxl') 

        df_zero.to_excel(writer, "Parcela Zero",index=False)  

        writer.save()

        book = load_workbook(os.path.join(*caminho_lst))

        writer_2 = pd.ExcelWriter(os.path.join(*caminho_lst), engine='openpyxl')

        writer_2.book = book

        writer_2.sheets = dict((ws.title, ws) for ws in book.worksheets)

        for chave in writer_2.sheets:
            if chave == "Planejamento Plurianual":                
                std=book[chave]                 
                book.remove(std)
                break
                
        writer_2.sheets = dict((ws.title, ws) for ws in book.worksheets)
            
        df_planejamento.to_excel(writer_2, "Planejamento Plurianual",index=False)  

        writer_2.save()


        retorno = "0"+","
        return retorno
    except:
        retorno = "1"+","+str(traceback.format_exc())
        return retorno



